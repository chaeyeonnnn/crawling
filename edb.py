# 크롤링해서 엑셀파일로 저장하는 함수

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
import urllib3.exceptions

from time import sleep
import logging
import datetime
import time
import openpyxl

from selenium import webdriver

logging.basicConfig(filename='error.log', level=logging.ERROR, format='%(asctime)s - %(message)s')

def get_data():
    new_data = []
    url = 'https://www.exploit-db.com/'
    #chrome_options = webdriver.ChromeOptions()
    #chrome_options.add_argument('--headless')  # 무허용 모드 활성화

    # Chrome WebDriver 초기화
    s = Service('/opt/homebrew/bin/chromedriver')
    driver = webdriver.Chrome(service=s)

    driver.get(url)
    driver.maximize_window()
    driver.implicitly_wait(time_to_wait=5)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["DATE", "TITLE", "EDB-ID", "CVE", "TYPE", "CODE", "CATEGORY", "CHECK"])
    due_date = 20230901

    while 1:
        try:
            rows = driver.find_elements(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[2]/div/div[2]/div/table/tbody/tr')
            for i in range(len(rows)):
                time.sleep(2)
                row = rows[i]
                columns = row.find_elements(By.TAG_NAME, 'td')
                date = columns[0].text
                title = columns[4].text
                platform = columns[6].text
                check = 'O' if platform == "Python" or platform == "payload" else ''
                date_time = int(date.replace("-",""))

                if due_date < date_time:
                    link = columns[4].find_element(By.TAG_NAME, 'a')
                    link.click()

                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6')))
                    #eb_id, cve, code는 title 누르고 들어가서 내용 따로 크롤링해서 다시 뒤로가기해서 다음 내용 뽑게

                    edb_id_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6')
                    edb_id = edb_id_element.text
                    cve_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[2]/h6')
                    cve = cve_element.text
                    type_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[2]/div/div[1]/div/div/div/div[2]/h6/a')
                    type = type_element.text
                    code_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[1]/div/pre/code')
                    code = code_element.text

                    try:
                        ws.append([date, title, edb_id, cve, type, code, platform, check])
                        driver.back()
                    except Exception:
                        driver.back()

                else:
                    print("finished")
                    wb.save("exploit_data.xlsx")
                    driver.quit()

            next_button = driver.find_element(By.XPATH, '//*[@id="exploits-table_next"]/a')
            if 'disabled' in next_button.get_attribute('class'):
                break
            else:
                next_button.click()
                time.sleep(3)
                #WebDriverWait(driver, 10).until(EC.staleness_of(rows[0]))

        except StaleElementReferenceException:
                driver.back()

        except urllib3.exceptions.MaxRetryError as e:
            logging.error(f'MaxRetryError: {e}')
            driver.back()


get_data()


'''

# 시작 날짜랑 마지막 날짜 정해서 하는
import requests
from bs4 import BeautifulSoup
import collections
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
import urllib3.exceptions

from time import sleep
import logging
import datetime
import time
import openpyxl

from selenium import webdriver

logging.basicConfig(filename='error.log', level=logging.ERROR, format='%(asctime)s - %(message)s')

def get_data():
    new_data = []
    url = 'https://www.exploit-db.com/'
    #chrome_options = webdriver.ChromeOptions()
    #chrome_options.add_argument('--headless')  # 무허용 모드 활성화

    # Chrome WebDriver 초기화
    s = Service('/opt/homebrew/bin/chromedriver')
    driver = webdriver.Chrome(service=s)

    driver.get(url)
    driver.maximize_window()
    driver.implicitly_wait(time_to_wait=5)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["DATE", "TITLE", "EDB-ID", "CVE", "TYPE", "CODE", "CATEGORY", "CHECK"])
    due_date = 20230824
    end_date = 20230904

    while 1:
        try:
            rows = driver.find_elements(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[2]/div/div[2]/div/table/tbody/tr')
            for i in range(len(rows)):
                time.sleep(2)
                row = rows[i]
                columns = row.find_elements(By.TAG_NAME, 'td')
                date = columns[0].text
                title = columns[4].text
                platform = columns[6].text
                check = 'O' if platform == "Python" or platform == "payload" else ''
                date_time = int(date.replace("-",""))

                if due_date <= date_time and date_time<=end_date:
                    link = columns[4].find_element(By.TAG_NAME, 'a')
                    link.click()

                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6')))
                    #eb_id, cve, code는 title 누르고 들어가서 내용 따로 크롤링해서 다시 뒤로가기해서 다음 내용 뽑게

                    edb_id_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6')
                    edb_id = edb_id_element.text
                    cve_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[2]/h6')
                    cve = cve_element.text
                    type_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[2]/div/div[1]/div/div/div/div[2]/h6/a')
                    type = type_element.text
                    code_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[1]/div/pre/code')
                    code = code_element.text

                    try:
                        ws.append([date, title, edb_id, cve, type, code, platform, check])
                        driver.back()
                    except Exception:
                        driver.back()
            
                elif date_time > end_date:
                    link = columns[4].find_element(By.TAG_NAME, 'a')
                    link.click()
                    driver.back()
                
                else:
                    print("finished")
                    wb.save("exploit_data.xlsx")
                    driver.quit()
                    break

            next_button = driver.find_element(By.XPATH, '//*[@id="exploits-table_next"]/a')
            if 'disabled' in next_button.get_attribute('class'):
                break
            else:
                next_button.click()
                time.sleep(3)
                #WebDriverWait(driver, 10).until(EC.staleness_of(rows[0]))

        except StaleElementReferenceException:
                driver.back()

        except urllib3.exceptions.MaxRetryError as e:
            logging.error(f'MaxRetryError: {e}')
            driver.back()


get_data()

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook 

url = 'https://www.exploit-db.com/'

s = Service('/opt/homebrew/bin/chromedriver')
driver = webdriver.Chrome(service=s)

driver.get(url)
driver.maximize_window()
driver.implicitly_wait(time_to_wait=5)

wb = Workbook() 
ws = wb.active 
ws.append(["DATE", "TITLE", "EDB-ID", "CVE", "TYPE", "CODE", "CATEGORY", "CHECK"])  # 헤더 추가

while True:
    rows = driver.find_elements(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[2]/div/div[2]/div/table/tbody/tr')

    for row in rows:
        columns = row.find_elements(By.TAG_NAME, 'td')
        date = columns[0].text
        title = columns[4].text
        platform = columns[6].text
        check = 'O' if platform == "Python" or platform == "payload" else ''

        link = columns[4].find_element(By.TAG_NAME, 'a')
        link.click()

        try:
            WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6')))
        except StaleElementReferenceException:
            pass

        edb_id = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6').text
        cve = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[2]/h6').text
        type = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[2]/div/div[1]/div/div/div/div[2]/h6/a').text
        code = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[1]/div/pre/code').text

        ws.append([date, title, edb_id, cve, type, code, platform, check])
        driver.back()

        data = [date, title, edb_id, cve, type, code, platform, check]
        ws.append(data) 

    next_button = driver.find_element(By.XPATH, '//*[@id="exploits-table_next"]/a')
    if 'disabled' in next_button.get_attribute('class'):
        break
    else:
        next_button.click()
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[2]/div/div[2]/div/table/tbody/tr')))
        
wb.save("exploit_data.xlsx")
driver.quit()

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

url = 'https://www.exploit-db.com/'

s = Service('/opt/homebrew/bin/chromedriver')
driver = webdriver.Chrome(service=s)

driver.get(url)
driver.maximize_window()
driver.implicitly_wait(time_to_wait=5)

print("DATE", "TITLE", "EDB-ID", "CVE", "TYPE", "CODE", "CATEGORY", "CHECK")

while True:
    rows = driver.find_elements(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[2]/div/div[2]/div/table/tbody/tr')

    for row in rows:
        columns = row.find_elements(By.TAG_NAME, 'td')
        date = columns[0].text
        title = columns[4].text
        platform = columns[6].text
        check = 'O' if platform == "Python" or platform == "payload" else ''

        link = columns[4].find_element(By.TAG_NAME, 'a')
        link.click()

        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6')))
        except TimeoutException:
            driver.back()
            continue

        edb_id_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6')
        edb_id = edb_id_element.text
        cve_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[2]/h6')
        cve=cve_element.text
        type_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[2]/div/div[1]/div/div/div/div[2]/h6/a')
        type=type_element.text
        code_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[1]/div/pre/code')
        code=code_element.text

        print(date, title, edb_id, cve, type, code, platform, check)
        driver.back()

    next_button = driver.find_element(By.XPATH, '//*[@id="exploits-table_next"]/a')
    if 'disabled' in next_button.get_attribute('class'):
        break
    else:
        next_button.click()
        try:
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[2]/div/div[2]/div/table/tbody/tr[1]')))
        except TimeoutException:
            break
driver.quit()


from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException
import time

url = 'https://www.exploit-db.com/'

s = Service('/opt/homebrew/bin/chromedriver')
driver = webdriver.Chrome(service=s)

driver.get(url)
driver.maximize_window()
driver.implicitly_wait(time_to_wait=5)

print("DATE", "TITLE", "EDB-ID", "CVE", "TYPE", "CODE", "CATEGORY", "CHECK")

while True:
    rows = driver.find_elements(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[2]/div/div[2]/div/table/tbody/tr')

    for i in range (len(rows)):
        try:
            row =driver.find_elements(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[2]/div/div[2]/div/table/tbody/tr')[i]
            columns = row.find_elements(By.TAG_NAME, 'td')
            date = columns[0].text
            title = columns[4].text
            platform = columns[6].text
            check = 'O' if platform == "Python" or platform == "payload" else ''

            link = columns[4].find_element(By.TAG_NAME, 'a')
            link.click()

            try:
                edb_id_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6')
                edb_id = edb_id_element.text
                cve_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[2]/h6')
                cve=cve_element.text
                type_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[2]/div/div[1]/div/div/div/div[2]/h6/a')
                type=type_element.text
                code_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[1]/div/pre/code')
                code=code_element.text
                print(date, title, edb_id, cve, type, code, platform, check)
                driver.back()
            except StaleElementReferenceException:
                driver.back()
                link = columns[4].find_element(By.TAG_NAME, 'a')
                link.click()

        except StaleElementReferenceException:
            driver.back()
            continue

    

    next_button = driver.find_element(By.XPATH, '//*[@id="exploits-table_next"]/a')
    if 'disabled' in next_button.get_attribute('class'):
        break
    else:
        next_button.click()
        #WebDriverWait(driver, 10).until(EC.staleness_of(rows[0]))
        time.sleep(2)

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException
import time

url = 'https://www.exploit-db.com/'

s = Service('/opt/homebrew/bin/chromedriver')
driver = webdriver.Chrome(service=s)

driver.get(url)
driver.maximize_window()
driver.implicitly_wait(time_to_wait=5)

print("DATE", "TITLE", "EDB-ID", "CVE", "TYPE", "CODE", "CATEGORY", "CHECK")

while True:
    try:
        rows = driver.find_elements(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[2]/div/div[2]/div/table/tbody/tr')

        for i in range(len(rows)):
            try:
                row = rows[i]
                columns = row.find_elements(By.TAG_NAME, 'td')
                date = columns[0].text
                title = columns[4].text
                platform = columns[6].text
                check = 'O' if platform == "Python" or platform == "payload" else ''

                link = columns[4].find_element(By.TAG_NAME, 'a')
                link.click()

                try:
                    edb_id_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6')
                    edb_id = edb_id_element.text
                    cve_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[2]/h6')
                    cve = cve_element.text
                    type_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[2]/div/div[1]/div/div/div/div[2]/h6/a')
                    type = type_element.text
                    code_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[1]/div/pre/code')
                    code = code_element.text
                    print(date, title, edb_id, cve, type, code, platform, check)
                    driver.back()
                except StaleElementReferenceException:
                    driver.back()
                    link = columns[4].find_element(By.TAG_NAME, 'a')
                    link.click()

            except StaleElementReferenceException:
                continue

        next_button = driver.find_element(By.XPATH, '//*[@id="exploits-table_next"]/a')
        if 'disabled' in next_button.get_attribute('class'):
            break
        else:
            driver.execute_script("arguments[0].scrollIntoView();", next_button)
            next_button.click()
            time.sleep(2)

    except StaleElementReferenceException:
        continue



from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

url = 'https://www.exploit-db.com/'

s = Service('/opt/homebrew/bin/chromedriver')
driver = webdriver.Chrome(service=s)

driver.get(url)
driver.maximize_window()
driver.implicitly_wait(time_to_wait=5)

print("DATE", "TITLE", "EDB-ID", "CVE", "TYPE", "CATEGORY", "CHECK")

while True:
    rows = driver.find_elements(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[2]/div/div[2]/div/table/tbody/tr')

    for row in rows:
        columns = row.find_elements(By.TAG_NAME, 'td')
        date = columns[0].text
        title = columns[4].text
        platform = columns[6].text
        check = 'O' if platform == "Python" or platform == "payload" else ''

        link = columns[4].find_element(By.TAG_NAME, 'a')
        link.click()

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6')))

        edb_id = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6').text
        cve = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[2]/h6').text
        type = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[2]/div/div[1]/div/div/div/div[2]/h6/a').text
        code = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[1]/div/pre/code').text

        print(date, title, edb_id, cve, type, platform, check)
        driver.back()

    next_button = driver.find_element(By.XPATH, '//*[@id="exploits-table_next"]/a')
    if 'disabled' in next_button.get_attribute('class'):
        break
    else:
        next_button.click()
        WebDriverWait(driver, 10).until(EC.staleness_of(rows[0]))

driver.quit()



from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import TimeoutException
import time
import openpyxl
import re
import logging 
from datetime import datetime


logging.basicConfig(filename='crawler.log', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')


def get_data():
    url = 'https://www.exploit-db.com/'
    s = Service('/opt/homebrew/bin/chromedriver')
    driver = webdriver.Chrome(service=s)
    
    driver.get(url)
    driver.maximize_window()
    driver.implicitly_wait(time_to_wait=5)

    # 엑셀 파일 만들고
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["DATE", "TITLE", "EDB-ID", "CVE", "TYPE", "CODE", "CATEGORY", "CHECK"])
    
    target_date = datetime(2023, 8, 1)


    while True:
        try:
            rows = driver.find_elements(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[2]/div/div[2]/div/table/tbody/tr')
            #여기에서 이제 나와있는 목록들 내용 다 뽑을 수 있음

            for i in range(len(rows)):
                row = rows[i]
                columns = row.find_elements(By.TAG_NAME, 'td')
                date = columns[0].text
                title = columns[4].text
                platform = columns[6].text
                check = 'O' if platform == "Python" or platform == "payload" else ''


                date_obj = datetime.strptime(date, "%Y-%m-%d")

                wait = WebDriverWait(driver, 20) 
                link = columns[4].find_element(By.TAG_NAME, 'a')
                #link = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="exploits-table"]/tbody/tr[1]/td[5]/a')))
                link.click()
                #eb_id, cve, code는 title 누르고 들어가서 내용 따로 크롤링해서 다시 뒤로가기해서 다음 내용 뽑게

                try:
                    edb_id_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6')
                    edb_id = edb_id_element.text
                    cve_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[2]/h6')
                    cve = cve_element.text
                    type_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[2]/div/div[1]/div/div/div/div[2]/h6/a')
                    type = type_element.text
                    code_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[1]/div/pre/code')
                    code = code_element.text
                    #자금 여기서 이상한 문자열 처리 문제 생김 !!!!!!!!!!!!!!!!!!!!!
                    try:
                        if date_obj >= target_date:
                        # 데이터를 엑셀에 추가해주기
                            ws.append([date, title, edb_id, cve, type, code, platform, check])
                            driver.back()

                    except StaleElementReferenceException as stale_exception:
                        logging.error(f'StaleElementReferenceException occurred: {stale_exception}')
                        driver.back()
                    
                    except Exception as e:
                        logging.error(f'Exception occurred: {e}')
                        driver.back()
                        
                        link = columns[4].find_element(By.TAG_NAME, 'a')
                        link.click()
                        break
                        
                    driver.back()

                except Exception as e:
                    driver.refresh()
                    time.sleep(2)
                    driver.back()
                    continue



            #next_button = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, '//*[@id="exploits-table_next"]/a')))
            next_button = driver.find_element(By.XPATH, '//*[@id="exploits-table_next"]/a')

            if 'disabled' in next_button.get_attribute('class'):
                break
            else:
                driver.execute_script("arguments[0].scrollIntoView();", next_button)
                next_button.click()
                time.sleep(2)
            #다음 페이지로 넘겨주다가 끝까지 가면 크롤링 그만하도록

        except Exception as e:
            continue

    # 엑셀 파일로 저장
    wb.save("exploit_data.xlsx")
    
if __name__ == "__main__":
    get_data()

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException
import logging


import time
import openpyxl
import re

logging.basicConfig(filename='crawler.log', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

def get_data():
    url = 'https://www.exploit-db.com/'
    s = Service('/opt/homebrew/bin/chromedriver')
    driver = webdriver.Chrome(service=s)
    
    driver.get(url)
    driver.maximize_window()
    driver.implicitly_wait(time_to_wait=5)

    # 엑셀 파일 만들고
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["DATE", "TITLE", "EDB-ID", "CVE", "TYPE", "CODE", "CATEGORY", "CHECK"])

    while True:
        try:
            rows = driver.find_elements(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[2]/div/div[2]/div/table/tbody/tr')
            #여기에서 이제 나와있는 목록들 내용 다 뽑을 수 있음

            for i in range(len(rows)):
                try:
                    row = rows[i]
                    columns = row.find_elements(By.TAG_NAME, 'td')
                    date = columns[0].text
                    title = columns[4].text
                    platform = columns[6].text
                    check = 'O' if platform == "Python" or platform == "payload" else ''

                    link = columns[4].find_element(By.TAG_NAME, 'a')
                    link.click()
                    #eb_id, cve, code는 title 누르고 들어가서 내용 따로 크롤링해서 다시 뒤로가기해서 다음 내용 뽑게

                    try:
                        edb_id_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6')
                        edb_id = edb_id_element.text
                        cve_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[2]/h6')
                        cve = cve_element.text
                        type_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[2]/div/div[1]/div/div/div/div[2]/h6/a')
                        type = type_element.text
                        code_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[1]/div/pre/code')
                        code = code_element.text
                        #자금 여기서 이상한 문자열 처리 문제 생김 !!!!!!!!!!!!!!!!!!!!!

                        # 데이터를 엑셀에 추가해주기
                        ws.append([date, title, edb_id, cve, type, code, platform, check])
                        driver.back()

                    except StaleElementReferenceException as stale_exception:
                        logging.error(f'StaleElementReferenceException occurred: {stale_exception}')
                        continue
                    except Exception as e:
                        logging.error(f'Exception occurred: {e}')
                        driver.back()
                        link = columns[4].find_element(By.TAG_NAME, 'a')
                        link.click()
                        break
                except StaleElementReferenceException:
                    continue
            
            #next_button = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, '//*[@id="exploits-table_next"]/a')))
            next_button = driver.find_element(By.XPATH, '//*[@id="exploits-table_next"]/a')

            if 'disabled' in next_button.get_attribute('class'):
                break
            else:
                driver.execute_script("arguments[0].scrollIntoView();", next_button)
                next_button.click()
                time.sleep(2)
            #다음 페이지로 넘겨주다가 끝까지 가면 크롤링 그만하도록

        except StaleElementReferenceException:
            continue
    # 엑셀 파일로 저장
    wb.save("exploit_data.xlsx")

if __name__ == "__main__":
    get_data()

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

url = 'https://www.exploit-db.com/'

s = Service('/opt/homebrew/bin/chromedriver')
driver = webdriver.Chrome(service=s)

driver.get(url)
driver.maximize_window()
driver.implicitly_wait(time_to_wait=5)

print("DATE", "TITLE", "EDB-ID", "CVE", "TYPE", "CODE", "CATEGORY", "CHECK")

while True:
    rows = driver.find_elements(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[2]/div/div[2]/div/table/tbody/tr')

    for row in rows:
        columns = row.find_elements(By.TAG_NAME, 'td')
        date = columns[0].text
        title = columns[4].text
        platform = columns[6].text
        check = 'O' if platform == "Python" or platform == "payload" else ''

        link = columns[4].find_element(By.TAG_NAME, 'a')
        link.click()

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6')))

        edb_id = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6').text
        cve = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[2]/h6').text
        type = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[2]/div/div[1]/div/div/div/div[2]/h6/a').text
        code = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[1]/div/pre/code').text

        print(date, title, edb_id, cve, type,code,  platform, check)
        driver.back()

    next_button = driver.find_element(By.XPATH, '//*[@id="exploits-table_next"]/a')
    if 'disabled' in next_button.get_attribute('class'):
        break
    else:
        next_button.click()
        WebDriverWait(driver, 10).until(EC.staleness_of(rows[0]))

driver.quit()






from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import NoSuchElementException

import logging
import datetime
import time
import openpyxl
import re


#logging.basicConfig(filename='crawler.log', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

def get_data():
    url = 'https://www.exploit-db.com/'
    s = Service('/opt/homebrew/bin/chromedriver')
    driver = webdriver.Chrome(service=s)
    
    driver.get(url)
    driver.maximize_window()
    driver.implicitly_wait(time_to_wait=5)

    current_date = datetime.datetime.now()

    # 엑셀 파일 만들고
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["DATE", "TITLE", "EDB-ID", "CVE", "TYPE", "CODE", "CATEGORY", "CHECK"])
    due_date = 20230801

    while True:
        try:
            rows = driver.find_elements(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[2]/div/div[2]/div/table/tbody/tr')
            #여기에서 이제 나와있는 목록들 내용 다 뽑을 수 있음

            for i in range(len(rows)):
                try:
                    row = rows[i]
                    columns = row.find_elements(By.TAG_NAME, 'td')
                    date = columns[0].text
                    title = columns[4].text
                    platform = columns[6].text
                    check = 'O' if platform == "Python" or platform == "payload" else ''
                    date_time = datetime.strptime(date, "%Y-%m-%d")
                    
                    link = columns[4].find_element(By.TAG_NAME, 'a')
                    link.click()

                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6')))

                    #eb_id, cve, code는 title 누르고 들어가서 내용 따로 크롤링해서 다시 뒤로가기해서 다음 내용 뽑게
                    edb_id_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6')
                    edb_id = edb_id_element.text
                    cve_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[2]/h6')
                    cve = cve_element.text
                    type_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[2]/div/div[1]/div/div/div/div[2]/h6/a')
                    type = type_element.text
                    code_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[1]/div/pre/code')
                    code = code_element.text

                    date_time = int(date.replace("-",""))

                    if date_time < due_date:
                        driver.close()
                        driver.switch_to.window(driver.window_handles[0]) 
                        break
                    else: 
                        try:
                            # 데이터를 엑셀에 추가해주기
                            ws.append([date, title, edb_id, cve, type, code, platform, check])
                            driver.close()  # 현재 창(웹 페이지) 닫기
                            driver.switch_to.window(driver.window_handles[0])   
                        except Exception as e:
                            logging.error(f'Exception occurred: {e}')
                            driver.close()  # 현재 창(웹 페이지) 닫기
                            driver.switch_to.window(driver.window_handles[0])  # 이전 창으로 전환
                            break
                except Exception:
                    continue

            try:
                next_button = driver.find_element(By.XPATH, '//*[@id="exploits-table_next"]/a')

                if 'disabled' in next_button.get_attribute('class'):
                    break
                else:
                    driver.execute_script("arguments[0].scrollIntoView();", next_button)
                    next_button.click()
                    time.sleep(2)
                #다음 페이지로 넘겨주다가 끝까지 가면 크롤링 그만하도록

            except NoSuchElementException:
                break

        # 엑셀 파일로 저장
            wb.save("exploit_data.xlsx")
        except Exception as e:
            print(e)
            # 예외 처리 코드 추가
            #logging.error(f'Exception occurred: {e}')

get_data()



from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

url = 'https://www.exploit-db.com/'

s = Service('/opt/homebrew/bin/chromedriver')
driver = webdriver.Chrome(service=s)

driver.get(url)
driver.maximize_window()
driver.implicitly_wait(time_to_wait=5)

print("DATE", "TITLE", "EDB-ID", "CVE", "TYPE", "CODE", "CATEGORY", "CHECK")

while True:
    rows = driver.find_elements(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[2]/div/div[2]/div/table/tbody/tr')

    for row in rows:
        columns = row.find_elements(By.TAG_NAME, 'td')
        date = columns[0].text
        title = columns[4].text
        platform = columns[6].text
        check = 'O' if platform == "Python" or platform == "payload" else ''

        link = columns[4].find_element(By.TAG_NAME, 'a')
        link.click()

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6')))

        edb_id = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6').text
        cve = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[2]/h6').text
        type = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[2]/div/div[1]/div/div/div/div[2]/h6/a').text
        code = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[1]/div/pre/code').text

        print(date, title, edb_id, cve, type,code,  platform, check)
        driver.back()

    next_button = driver.find_element(By.XPATH, '//*[@id="exploits-table_next"]/a')
    if 'disabled' in next_button.get_attribute('class'):
        break
    else:
        next_button.click()
        WebDriverWait(driver, 10).until(EC.staleness_of(rows[0]))

driver.quit()
'''
