# 기존에 저장되어있던 파일과 실시간으로 크롤링한 데이터를 비교하여 업데이트된 데이터만 따로 파일로 새로 저장하는 코드 

import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
import time
from selenium.common.exceptions import StaleElementReferenceException
import urllib3.exceptions
import logging

# 매주 수정된 부분 있으면 그 부분 알려줘야됨
# 이전에 저장한 엑셀 파일을 열어서 리스트에 넣음
def previous():
    file_path = 'exploit_data.xlsx'
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    previous_data = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        previous_data.append(row)
    print(previous_data)
    return previous_data


# 업데이트된 데이터 실시간으로 크롤링해서 리스트에 넣어놓고
def recent_data():
    new_data = [] 
    url = 'https://www.exploit-db.com/'

    s = Service('/opt/homebrew/bin/chromedriver')
    driver = webdriver.Chrome(service=s)

    driver.get(url)
    driver.maximize_window()
    driver.implicitly_wait(time_to_wait=5)

    due_date = 20230824

    while True:
        try:
            rows = driver.find_elements(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[2]/div/div[2]/div/table/tbody/tr')

            for i in range(len(rows)):
                time.sleep(3)
                row = rows[i]
                columns = row.find_elements(By.TAG_NAME, 'td')
                date = columns[0].text
                title = columns[4].text
                platform = columns[6].text
                check = 'O' if platform == "Python" or platform == "payload" else ''
                date_time = int(date.replace("-",""))

                if due_date < date_time:
                    link = columns[4].find_element(By.TAG_NAME, 'a')
                    link.click()
                    time.sleep(3)

                    #eb_id, cve, code는 title 누르고 들어가서 내용 따로 크롤링해서 다시 뒤로가기해서 다음 내용 뽑게

                    edb_id_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6')
                    edb_id = edb_id_element.text
                    cve_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[2]/h6')
                    cve = cve_element.text
                    type_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[2]/div/div[1]/div/div/div/div[2]/h6/a')
                    type = type_element.text
                    code_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[1]/div/pre/code')
                    code = code_element.text
                
                    try:
                        new_data.append([date, title, edb_id, cve, type, code, platform, check])
                        driver.back()
                    except Exception:
                        driver.back()
                else:
                    print("new data 다 스크래핑 완료") #--> 여기도 지금 출력됨, recent_data()까진 실행되는중
                    break


            next_button = driver.find_element(By.XPATH, '//*[@id="exploits-table_next"]/a')
            if 'disabled' in next_button.get_attribute('class'):
               #driver.close()
               #break
               driver.quit()
            else:
                next_button.click()
                time.sleep(3)
                #WebDriverWait(driver, 10).until(EC.staleness_of(rows[0]))

        except StaleElementReferenceException:
            driver.back()

        except urllib3.exceptions.MaxRetryError as e:
            logging.error(f'MaxRetryError: {e}')
            driver.back()

        return new_data

#new_data까지 출력 잘 됨 지금

def same_cve(new_data, previous_data):
    for i in range(len(new_data)):
        if i == 3:  
            continue
        if new_data[i] != previous_data[i]:
            return True
    return False

#바뀐 데이터만 새로 엑셀파일에 저장
def save_updated_data(previous_data, new_data):
    updated_data = []
    last_date = previous_data[-1][0] if previous_data else None
    
    for new_date in new_data:
        date = new_date[0]
        date_time = int(date.replace("-", ""))
        cve = new_date[3] 
        if last_date is None or date_time >= int(last_date.replace("-", "")) or same_cve(new_date, previous_data):
            if all(new_date[1:] != prev_data[1:] for prev_data in previous_data):
                updated_data.append(new_date)
    
    if updated_data:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["DATE", "TITLE", "EDB-ID", "CVE", "TYPE", "CODE", "CATEGORY", "CHECK"])

        for row in updated_data:
            ws.append(row)
        wb.save("updated_data.xlsx")
        print("New updated data" )
    else:
        print("No updated data")
        
save_updated_data(previous(), recent_data())

'''
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
import time
from selenium.common.exceptions import StaleElementReferenceException
import urllib3.exceptions
import logging

# 매주 수정된 부분 있으면 그 부분 알려줘야됨
# 이전에 저장한 엑셀 파일을 열어서 리스트에 넣음
def previous():
    file_path = 'exploit_data.xlsx'
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    previous_data = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        previous_data.append(row)
    print(previous_data)
    return previous_data


# 업데이트된 데이터 실시간으로 크롤링해서 리스트에 넣어놓고
def recent_data():
    new_data = [] 
    url = 'https://www.exploit-db.com/'

    s = Service('/opt/homebrew/bin/chromedriver')
    driver = webdriver.Chrome(service=s)

    driver.get(url)
    driver.maximize_window()
    driver.implicitly_wait(time_to_wait=5)

    due_date = 20230824

    while True:
        try:
            rows = driver.find_elements(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[2]/div/div[2]/div/table/tbody/tr')

            for i in range(len(rows)):
                time.sleep(3)
                row = rows[i]
                columns = row.find_elements(By.TAG_NAME, 'td')
                date = columns[0].text
                title = columns[4].text
                platform = columns[6].text
                check = 'O' if platform == "Python" or platform == "payload" else ''
                date_time = int(date.replace("-",""))

                if due_date < date_time:
                    link = columns[4].find_element(By.TAG_NAME, 'a')
                    link.click()
                    time.sleep(3)

                    #eb_id, cve, code는 title 누르고 들어가서 내용 따로 크롤링해서 다시 뒤로가기해서 다음 내용 뽑게

                    edb_id_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6')
                    edb_id = edb_id_element.text
                    cve_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[2]/h6')
                    cve = cve_element.text
                    type_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[2]/div/div[1]/div/div/div/div[2]/h6/a')
                    type = type_element.text
                    code_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[1]/div/pre/code')
                    code = code_element.text
                
                    try:
                        new_data.append([date, title, edb_id, cve, type, code, platform, check])
                        driver.back()
                    except Exception:
                        driver.back()
                else:
                    print("new data 다 스크래핑 완료") #--> 여기도 지금 출력됨, recent_data()까진 실행되는중
                    break


            next_button = driver.find_element(By.XPATH, '//*[@id="exploits-table_next"]/a')
            if 'disabled' in next_button.get_attribute('class'):
               #driver.close()
               #break
               driver.quit()
            else:
                next_button.click()
                time.sleep(3)
                #WebDriverWait(driver, 10).until(EC.staleness_of(rows[0]))

        except StaleElementReferenceException:
            driver.back()

        except urllib3.exceptions.MaxRetryError as e:
            logging.error(f'MaxRetryError: {e}')
            driver.back()

        return new_data

#new_data까지 출력 잘 됨 지금


#바뀐 데이터만 새로 엑셀파일에 저장
def save_updated_data(previous_data, new_data):
    updated_data = []
    last_date = previous_data[-1][0] if previous_data else None

    for new_date in new_data:
        date = new_date[0]
        date_time = int(date.replace("-", ""))
        if last_date is None or date_time >= int(last_date.replace("-", "")):
            if all(new_date[1:] != prev_data[1:] for prev_data in previous_data):
                updated_data.append(new_date)

    if updated_data:
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(["DATE", "TITLE", "EDB-ID", "CVE", "TYPE", "CODE", "CATEGORY", "CHECK"])

        for row in updated_data:
            ws.append(row)
        
        wb.save("updated_data.xlsx")
        print("New updated data" )

    else:
        print("No updated data")

#s = Service('/opt/homebrew/bin/chromedriver')
#driver = webdriver.Chrome(service=s)

save_updated_data(previous(), recent_data())


import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
import time
from selenium.common.exceptions import StaleElementReferenceException
import urllib3.exceptions
import logging

# 매주 수정된 부분 있으면 그 부분 알려줘야됨
# 이전에 저장한 엑셀 파일을 열어서 리스트에 넣음
def previous_data(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    previous_data = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        previous_data.append(row)
    print(previous_data)
    return previous_data


# 업데이트된 데이터 실시간으로 크롤링해서 리스트에 넣어놓고
def recent_data():
    new_data = [] 
    url = 'https://www.exploit-db.com/'

    s = Service('/opt/homebrew/bin/chromedriver')
    driver = webdriver.Chrome(service=s)

    driver.get(url)
    driver.maximize_window()
    driver.implicitly_wait(time_to_wait=5)

    due_date = 20230901

    while 1:
        try:
            rows = driver.find_elements(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[2]/div/div[2]/div/table/tbody/tr')

            for i in range(len(rows)):
                row = rows[i]
                columns = row.find_elements(By.TAG_NAME, 'td')
                date = columns[0].text
                title = columns[4].text
                platform = columns[6].text
                check = 'O' if platform == "Python" or platform == "payload" else ''
                date_time = int(date.replace("-",""))

                if due_date < date_time:
                    link = columns[4].find_element(By.TAG_NAME, 'a')
                    link.click()

                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6')))

                    #eb_id, cve, code는 title 누르고 들어가서 내용 따로 크롤링해서 다시 뒤로가기해서 다음 내용 뽑게

                    edb_id_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[1]/h6')
                    edb_id = edb_id_element.text
                    cve_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/div/div/div/div[2]/h6')
                    cve = cve_element.text
                    type_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div[2]/div/div[1]/div/div/div/div[2]/h6/a')
                    type = type_element.text
                    code_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[2]/div/div/div[2]/div[1]/div/pre/code')
                    code = code_element.text
                
                    try:
                        new_data.append([date, title, edb_id, cve, type, code, platform, check])
                        driver.back()

                    except Exception as e:
                        driver.back()
                else:
                    driver.quit()

            next_button = driver.find_element(By.XPATH, '//*[@id="exploits-table_next"]/a')
            if 'disabled' in next_button.get_attribute('class'):
                break
            else:
                next_button.click()
                time.sleep(3)
                #WebDriverWait(driver, 10).until(EC.staleness_of(rows[0]))

        except StaleElementReferenceException:
                driver.back()

        except urllib3.exceptions.MaxRetryError as e:
            logging.error(f'MaxRetryError: {e}')
            driver.back()

    driver.quit()
    return new_data


# 기존에 저장된 파일이랑 새로 크롤링한거 비교
def compare(previous_data, new_data):
    updated_data=[]
    for new_row in new_data:
        found = False
        for prev_row in previous_data:
            if new_row == prev_row:  # 코드 열 비교
                found = True
                continue
        if not found:
            updated_data.append(new_row)
    return updated_data


#바뀐 데이터만 새로 엑셀파일에 저장
def save_updated_data(previous_data, new_data):
    updated_data = []
    last_date = previous_data[-1][0] if previous_data else None

    for new_row in new_data:
        date = new_row[0]
        date_time = int(date.replace("-", ""))
        if last_date is None or date_time > int(last_date.replace("-", "")):
            updated_data.append(new_row)

    if updated_data:
        wb = openpyxl.load_workbook("updated_data.xlsx")
        ws = wb.active
        ws.append(["DATE", "TITLE", "EDB-ID", "CVE", "TYPE", "CODE", "CATEGORY", "CHECK"])

        for row in updated_data:
            ws.append(row)

        wb.save("updated_data.xlsx")
        print("New updated data" )

    else:
        print("No updated data")

data = 'exploit_data.xlsx'
previous_data_list = previous_data(data)
new_data_list = recent_data()
updated_data_list = compare(previous_data_list, new_data_list)
save_updated_data(updated_data_list)
'''
