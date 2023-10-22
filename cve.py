
import json
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import openpyxl
from selenium.common.exceptions import NoSuchElementException
import requests

driver = webdriver.Chrome()
url = 'https://github.com/CVEProject/cvelistV5/tree/main/cves'
driver.get(url)
driver.maximize_window()
time.sleep(5)

wb = openpyxl.Workbook()
ws = wb.active
ws.append(["파일 이름", '파일 내용'])

rows = driver.find_elements(By.TAG_NAME, 'tr')
num_rows = len(rows)

for year_i in range(1, num_rows):
    year_path = f'//*[@id="folder-row-{year_i}"]'
    time.sleep(3)
    year_row = driver.find_element(By.XPATH, year_path)
    columns = year_row.find_elements(By.CLASS_NAME, 'react-directory-commit-age')
    try:
        for column in columns:
            last_commit_date = column.text
            parts = last_commit_date.split()

            if parts[0].isdigit():
                day = int(parts[0])
                unit = parts[1]
            else:
                continue

            if (day <= 6 and unit == "days") or unit == "hours" or unit == "minutes" or last_commit_date == "yesterday":
                year_row.click()
                print(f"{year_i + 1998} 누름")
                time.sleep(5)

                file_links = driver.find_elements(By.XPATH, '//*[@href]')
                for link in file_links:
                    file_url = link.get_attribute('href')
                    if file_url.endswith('.json'):

                        # 파일 다운로드
                        response = requests.get(file_url)
                        filename = file_url.split('/')[-1]
                        with open(filename, 'wb') as file:
                            file.write(response.content)

                        with open(filename, 'r', encoding='utf-8') as json_file:
                            json_data = json.load(json_file)
                            content = json.dumps(json_data, indent=4)
                            ws.append([filename, content])

                driver.back()

            else:
                continue
        wb.save("updated_cve_data.xlsx")

    except NoSuchElementException:
        driver.quit()
        break

driver.quit()

'''

import json
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import openpyxl
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


driver = webdriver.Chrome()
url = 'https://github.com/CVEProject/cvelistV5/tree/main/cves'
driver.get(url)
driver.maximize_window()
time.sleep(5)

wb = openpyxl.Workbook()
ws = wb.active
ws.append(["파일이름",'파일 내용'])
   

rows = driver.find_elements(By.TAG_NAME, 'tr')
num_rows = len(rows)

#이게 연도 for문 ,, 제일 대빵
for year_i in range(1, num_rows):
    year_path = f'//*[@id="folder-row-{year_i}"]'
    time.sleep(3)
    year_row = driver.find_element(By.XPATH, year_path)
    columns = year_row.find_elements(By.CLASS_NAME, 'react-directory-commit-age')
    try:
        # 연도 선택
        for column in columns:
            last_commit_date = column.text
            parts = last_commit_date.split()

            if parts[0].isdigit():
                day = int(parts[0])
                unit = parts[1]
            else:
                continue

            if (day <= 6 and unit == "days") or unit == "hours" or unit == "minutes" or last_commit_date == "yesterday":
                year_row.click()
                print(f"{year_i + 1998} 누름")
                time.sleep(5)

                update_names = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'Link--primary.Truncate-text')))
                update_datas = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'js-file-content.Details-content--hidden.position-relative'))
        )
                '''
                #csv로 올릴때 따로 변환?
                for update_name, update_data in zip(update_names, update_datas):
                    json_name = json.loads(update_name.get_attribute('data-ga-click-data'))
                    json_data = json.loads(update_data.get_attribute('data-ga-click-data'))

                    name_values = [list(name.values()) for name in json_name]
                    data_values = [list(data.values()) for data in json_data]

                    for name, data in zip(name_values, data_values):
                        row = name + data
                        ws.append(row)
                '''
                # last commit message눌려서 +- 자료 text로 그냥 긁어서 zip으로 해서 같이 저장하는 방법
                for update_name, update_data in zip(update_names, update_datas):
                    ws.append([update_name.text,update_data.text])
                    print(update_name.text,update_data.text)
                driver.back()

            else:
                continue
        wb.save("updated_cve_data.xlsx") 

    except NoSuchElementException:
        driver.quit() 
        break



from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import os
import csv
import logging
from selenium.common.exceptions import NoSuchElementException


logging.basicConfig(filename='cve_error.log', level=logging.ERROR, format='%(asctime)s - %(message)s')

csv_file_name = 'cve_data.csv'

with open(csv_file_name, 'w', newline='', encoding='utf-8-sig') as csvfile:
    csv_writer = csv.writer(csvfile)
    csv_writer.writerow(['파일 이름', '파일 내용'])

if not os.path.exists('cve_data_files'):
    os.makedirs('cve_data_files')

driver = webdriver.Chrome()

url = 'https://github.com/CVEProject/cvelistV5/tree/main/cves'

driver.get(url)
driver.maximize_window()
time.sleep(5)

rows= driver.find_elements(By.TAG_NAME,'tr')
num_rows = len(rows)
#행의 개수

#몇년에 들어감
for year_i in range(num_rows):
    year_path = f'//*[@id="folder-row-{year_i}"]'
    year_row = driver.find_element(By.XPATH, year_path)
    columns = year_row.find_elements(By.CLASS_NAME,'react-directory-commit-age')
    for column in columns:
        last_commit_date = column.text
        parts = last_commit_date.split()

        if parts[0].isdigit():
            day = int(parts[0])
            unit = parts[1]
        else:
            year_i += 1
            continue

        if (day <= 6 and unit == "days") or unit == "hours" or unit == "minutes" or last_commit_date == "yesterday":
            year_row.click()
            time.sleep(2)

            for xxx_i in range(num_rows):
                year_path = f'//*[@id="folder-row-{xxx_i}"]'
                year_row = driver.find_element(By.XPATH, year_path)
                columns = year_row.find_elements(By.CLASS_NAME,'react-directory-commit-age')
                for column in columns:
                    last_commit_date = column.text
                    parts = last_commit_date.split()

                    if parts[0].isdigit():
                        day = int(parts[0])
                        unit = parts[1]
                    else:
                        xxx_i += 1
                        continue

                    if (day <= 6 and unit == "days") or unit == "hours" or unit == "minutes" or last_commit_date == "yesterday":
                        year_row.click()
                        time.sleep(2)
                        
                        #~xxx의 json 파일들 훑어서 최근에 업데이트된 정보에 또 들어가
                        for json_i in range(num_rows):
                            year_path = f'//*[@id="folder-row-{json_i}"]'
                            year_row = driver.find_element(By.XPATH, year_path)
                            columns = year_row.find_elements(By.CLASS_NAME,'react-directory-commit-age')
                            for column in columns:
                                last_commit_date = column.text
                                parts = last_commit_date.split()

                                if parts[0].isdigit():
                                    day = int(parts[0])
                                    unit = parts[1]
                                else:
                                    json_i += 1
                                    continue

                                if (day <= 6 and unit == "days") or unit == "hours" or unit == "minutes" or last_commit_date == "yesterday":
                                    year_row.click()
                                    time.sleep(2)
                                    download_button_xpath = '//*[@id="repos-sticky-header"]/div[1]/div[2]/div[2]/div[1]/div[1]/span/button/svg'
                                    download_button = driver.find_element(By.XPATH, download_button_xpath)
                                    download_button.click()

                                    time.sleep(5)

                                    downloaded_content = driver.page_source
                                    file_name = f'download_data_{year_i}_{xxx_i}_{json_i}.csv'

                                    with open(file_name, 'w', encoding='utf-8-sig') as csvfile:
                                        csv_writer = csv.writer(csvfile)
                                        csv_writer.writerow([file_name, downloaded_content])
                                    logging.info(f"파일 저장 완료: {file_name}")
                                    print(f"{file_name}을 저장했습니다.")

                                else:
                                    driver.back()
                                    continue
                    else:
                        driver.back()
                        continue
        else:
            driver.back()
            continue

from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import os
import csv
import logging
from selenium.common.exceptions import NoSuchElementException

logging.basicConfig(filename='cve_error.log', level=logging.ERROR, format='%(asctime)s - %(message)s')

csv_file_name = 'cve_data.csv'

with open(csv_file_name, 'w', newline='', encoding='utf-8-sig') as csvfile:
    csv_writer = csv.writer(csvfile)
    csv_writer.writerow(['파일 이름', '파일 내용'])

if not os.path.exists('cve_data_files'):
    os.makedirs('cve_data_files')

driver = webdriver.Chrome()

url = 'https://github.com/CVEProject/cvelistV5/tree/main/cves'

driver.get(url)
driver.maximize_window()
time.sleep(5)

def download_file(row, year_i, xxx_i, json_i):
    row.click()
    time.sleep(2)
    download_button_xpath = '//*[@id="repos-sticky-header"]/div[1]/div[2]/div[2]/div[1]/div[1]/span/button/svg'
    download_button = driver.find_element(By.XPATH, download_button_xpath)
    download_button.click()
    time.sleep(5)
    downloaded_content = driver.page_source
    file_name = f'download_data_{year_i}_{xxx_i}_{json_i}.csv'

    with open(file_name, 'w', encoding='utf-8-sig') as csvfile:
        csv_writer = csv.writer(csvfile)
        csv_writer.writerow([file_name, downloaded_content])
    logging.info(f"파일 저장 완료: {file_name}")
    print(f"{file_name}을 저장했습니다.")
    driver.back()

rows = driver.find_elements(By.TAG_NAME, 'tr')
num_rows = len(rows)

# Year 루프
for year_i in range(num_rows):
    year_path = f'//*[@id="folder-row-{year_i}"]'
    year_row = driver.find_element(By.XPATH, year_path)
    last_commit_date = year_row.find_element(By.CLASS_NAME, 'react-directory-commit-age').text
    parts = last_commit_date.split()

    if parts[0].isdigit():
        day = int(parts[0])
        unit = parts[1]
    else:
        year_i += 1
        continue

    if (day <= 6 and unit == "days") or unit == "hours" or unit == "minutes" or last_commit_date == "yesterday":
        # xxx 루프
        for xxx_i in range(num_rows):
            xxx_row = driver.find_element(By.XPATH, f'//*[@id="folder-row-{xxx_i}"]')
            last_commit_date = xxx_row.find_element(By.CLASS_NAME, 'react-directory-commit-age').text
            parts = last_commit_date.split()
            
            if parts[0].isdigit():
                day = int(parts[0])
                unit = parts[1]
            else:
                xxx_i += 1
                continue
            
            if (day <= 6 and unit == "days") or unit == "hours" or unit == "minutes" or last_commit_date == "yesterday":
                # json 루프
                for json_i in range(num_rows):
                    json_row = driver.find_element(By.XPATH, f'//*[@id="folder-row-{json_i}"]')
                    last_commit_date = json_row.find_element(By.CLASS_NAME, 'react-directory-commit-age').text
                    parts = last_commit_date.split()

                    if parts[0].isdigit():
                        day = int(parts[0])
                        unit = parts[1]
                    else:
                        json_i += 1
                        continue

                    if (day <= 6 and unit == "days") or unit == "hours" or unit == "minutes" or last_commit_date == "yesterday":
                        download_file(json_row, year_i, xxx_i, json_i)
                    else:
                        driver.back()
            else:
                driver.back()
    else:
        driver.back()

from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import os
import csv
import logging
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC



logging.basicConfig(filename='cve_error.log', level=logging.ERROR, format='%(asctime)s - %(message)s')

csv_file_name = 'cve_data.csv'

with open(csv_file_name, 'w', newline='', encoding='utf-8-sig') as csvfile:
    csv_writer = csv.writer(csvfile)
    csv_writer.writerow(['파일 이름', '파일 내용'])

if not os.path.exists('cve_data_files'):
    os.makedirs('cve_data_files')

driver = webdriver.Chrome()

url = 'https://github.com/CVEProject/cvelistV5/tree/main/cves'

driver.get(url)
driver.maximize_window()
time.sleep(5)

rows= driver.find_elements(By.TAG_NAME,'tr')
num_rows = len(rows)
#행의 개수

#몇년에 들어감
for year_i in range(num_rows):
    year_path = f'//*[@id="folder-row-{year_i}"]'
    #year_button = f'//*[@id="folder-row-{year_i}"]/td[2]/div'
    #year_button_row = driver.find_element(By.XPATH, year_button)

    try:
        time.sleep(3)
        year_row = driver.find_element(By.XPATH, year_path)
    except NoSuchElementException:
        break
    columns = year_row.find_elements(By.CLASS_NAME,'react-directory-commit-age')

    #여기가 지금 년도별 last commit date 다 쫘르륵,, 
    for column in columns:
        last_commit_date = column.text
        parts = last_commit_date.split()

        if parts[0].isdigit():
            day = int(parts[0])
            unit = parts[1]
        else:
            continue

        if (day <= 6 and unit == "days") or unit == "hours" or unit == "minutes" or last_commit_date == "yesterday":
            year_row.click()
            print(f"Clicked on year {year_i}")

            time.sleep(30)

            xxx_rows= driver.find_elements(By.TAG_NAME,'tr')
            xxx_num_rows = len(xxx_rows)
            for xxx_i in range(xxx_num_rows):
                xxx_path = f'//*[@id="folder-row-{xxx_i}"]'
                try:
                    xxx_row = driver.find_element(By.XPATH, xxx_path)
                except NoSuchElementException:
                    continue
                columns = xxx_row.find_elements(By.CLASS_NAME,'react-directory-commit-age')
                for column in columns:
                    last_commit_date = column.text
                    parts = last_commit_date.split()

                    if parts[0].isdigit():
                        day = int(parts[0])
                        unit = parts[1]
                    else:
                        continue

                    if (day <= 6 and unit == "days") or unit == "hours" or unit == "minutes" or last_commit_date == "yesterday":
                        xxx_row.click()
                        print(f"Clicked on xxx {xxx_i}")

                        time.sleep(10)

                        json_rows= driver.find_elements(By.TAG_NAME,'tr')
                        json_num_rows = len(json_rows)

                        #~xxx의 json 파일들 훑어서 최근에 업데이트된 정보에 또 들어가
                        for json_i in range(json_num_rows):
                            json_path = f'//*[@id="folder-row-{json_i}"]'
                            try:
                                json_row = driver.find_element(By.XPATH, json_path)
                            except NoSuchElementException:
                                continue
                            columns = json_row.find_elements(By.CLASS_NAME,'react-directory-commit-age')
                            for column in columns:
                                last_commit_date = column.text
                                parts = last_commit_date.split()

                                if parts[0].isdigit():
                                    day = int(parts[0])
                                    unit = parts[1]
                                else:
                                    continue

                                if (day <= 6 and unit == "days") or unit == "hours" or unit == "minutes" or last_commit_date == "yesterday":
                                    json_row.click()
                                    print(f"Clicked on json {json_i}")

                                    time.sleep(2)
                                    download_button_xpath = '//*[@id="repos-sticky-header"]/div[1]/div[2]/div[2]/div[1]/div[1]/span/button/svg'
                                    download_button = driver.find_element(By.XPATH, download_button_xpath)
                                    download_button.click()

                                    time.sleep(5)

                                    downloaded_content = driver.page_source
                                    file_name = f'download_data_{year_i}_{xxx_i}_{json_i}.csv'

                                    with open(file_name, 'w', encoding='utf-8-sig') as csvfile:
                                        csv_writer = csv.writer(csvfile)
                                        csv_writer.writerow([file_name, downloaded_content])
                                    logging.info(f"파일 저장 완료: {file_name}")
                                    print(f"{file_name}을 저장했습니다.")

                                else:
                                    continue
                    else:
                        continue
        else:
            continue
driver.quit()

from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import os
import csv
import logging
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import json

logging.basicConfig(filename='cve_error.log', level=logging.ERROR, format='%(asctime)s - %(message)s')

csv_file_name = 'new_cve_data.csv'

with open(csv_file_name, 'w', newline='', encoding='utf-8-sig') as csvfile:
    csv_writer = csv.writer(csvfile)
    csv_writer.writerow(['파일 이름', '파일 내용'])


driver = webdriver.Chrome()

url = 'https://github.com/CVEProject/cvelistV5/tree/main/cves'

driver.get(url)
driver.maximize_window()
time.sleep(5)

rows = driver.find_elements(By.TAG_NAME, 'tr')
num_rows = len(rows)

#이게 연도 for문 ,, 제일 대빵
for year_i in range(1,num_rows):
    year_path = f'//*[@id="folder-row-{year_i}"]'
    time.sleep(3)
    year_row = driver.find_element(By.XPATH, year_path)
    columns = year_row.find_elements(By.CLASS_NAME, 'react-directory-commit-age')
    
    #여기에서 이제 날짜 보고 클릭할 연도 고르기
    for column in columns: #column이 날짜야
        last_commit_date = column.text
        parts = last_commit_date.split()

        if parts[0].isdigit():
            day = int(parts[0])
            unit = parts[1]
        else:
            continue

        if (day <= 6 and unit == "days") or unit == "hours" or unit == "minutes" or last_commit_date == "yesterday":
            year_row.click()
            print(f"{year_i+1998} 누름!!!!!!!!!!!!!")
            time.sleep(10)
            
            #pdate_names = driver.find_elements(By.CLASS_NAME, 'Link--primary.Truncate-text')
            #update_datas = driver.find_elements(By.CLASS_NAME, 'js-file-content Details-content--hidden position-relative')
            # name = update_name.get_attribute('data_blob')
              #  if name:
               #     name_object = json.loads(name)
            
            update_names = driver.find_elements(By.CLASS_NAME, 'Link--primary.Truncate-text')
            update_datas = driver.find_elements(By.CLASS_NAME, 'js-file-content Details-content--hidden position-relative')


            for update_name, update_data in zip(update_names, update_datas):
                json_name = json.loads(update_name.get_attribute('data-ga-click-data'))
                json_data = json.loads(update_data.get_attribute('data-ga-click-data'))

                header = list(json_name[0].keys()) + list(json_data[0].keys())
                
                with open(csv_file_name, 'a', encoding='utf-8-sig', newline='') as csvfile:
                    csv_writer = csv.writer(csvfile)
                    csv_writer.writerow(header)

                    for name, data in zip(json_name, json_data):
                        row = list(name.values()) + list(data.values())
                        csv_writer.writerow(row)
                
                driver.back()
            
        else:
            continue

driver.quit()

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
import time
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import openpyxl
import os
import csv
import requests
import logging

logging.basicConfig(filename='cve_error.log', level=logging.ERROR, format='%(asctime)s - %(message)s')

csv_file_name = 'cve_data.csv'

with open(csv_file_name, 'w', newline='', encoding='utf-8-sig') as csvfile:
    csv_writer = csv.writer(csvfile)
    csv_writer.writerow(['File Name', 'File Content'])  
# 파일 토대 잡기, 이 밑에 row로 데이터 추가될 예정

if not os.path.exists('cve_data_files'):
    os.makedirs('cve_data_files')

s = Service('/opt/homebrew/bin/chromedriver')
driver = webdriver.Chrome(service=s)

url = 'https://github.com/CVEProject/cvelistV5/tree/main/cves'

driver.get(url)
driver.maximize_window()
time.sleep(5)

year_element = driver.find_element(By.XPATH, '//*[@id="folder-row-25"]/td[2]/div/div/h3/div/a')
year_element.click()

try:
    add_element = driver.find_element(By.XPATH, '//*[@id="folder-row-1"]/td[2]/div/div/h3/div/a')
    add_element.click()

except:
    logging.error("에러메시지")
    add_element = driver.find_element(By.XPATH, '//*[@id="folder-row-1"]/td[2]/div/div/h3/div/a')  # 다시 찾기
    add_element.click()

#여기까지 하면 0xxx에 까지 들어간거임


i = 1
while True:
    try:
        path = """//*[@id="repo-content-pjax-container"]/react-app/div/div/div[1]/div/div/main/div[2]/div/div[3]/div[3]/div/div/table"""
        # 이게 0xxxx들어가면 쫙 나오는 목록 표 xpath
     
        rows = driver.find_elements(By.XPATH,path)
        if i <len(rows):
            row = rows[i]
            link_element = row.find_element(By.TAG_NAME, 'a')
            link_element.click()
            time.sleep(2)

            download_link = driver.find_element(By.XPATH, '//*[@id="repos-sticky-header"]/div[1]/div[2]/div[2]/div[1]/div[1]/span/a')
            file_url = download_link.get_attribute('href')
            response = requests.get(file_url)

            file_name = 'CVE_data_' + str(i) + '.csv'
            file_content = response.text
            
            with open(os.path.join('cve_data_files', file_name), 'w', newline='', encoding='utf-8-sig') as file:
                file.write(file_content)
            logging.info(f"파일 저장 완료: {file_name}")


            with open(csv_file_name, 'a', newline='', encoding='utf-8-sig') as csvfile:
                csv_writer = csv.writer(csvfile)
                csv_writer.writerow([file_name, file_content])

            print(f"save {file_name}")
            
            driver.back()
            time.sleep(2)
            i += 1
        else:
            break

    except NoSuchElementException:
        break 

driver.quit()


import requests
import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import StaleElementReferenceException
import pandas as pd
import json

# Chrome 웹 드라이버를 실행
s = Service('/opt/homebrew/bin/chromedriver')
driver = webdriver.Chrome(service=s)
driver.maximize_window()
time.sleep(10)

# GitHub API 엔드포인트 및 토큰 설정
base_url = 'https://api.github.com/repos/CVEProject/cvelistV5/contents/cves'
token = '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'

# 다운로드할 연도 설정
year = 2023

# 결과를 저장할 디렉토리 생성
if not os.path.exists('CVE_files'):
    os.makedirs('CVE_files')

year_url = f'{base_url}{year:04d}'
headers = {'Authorization': f'token {token}'}

# 해당 연도의 폴더 내 파일 목록 가져오기
response = requests.get(year_url, headers=headers)

try:
    data = response.json()
    result_df = pd.DataFrame(columns=['File Name', 'Code'])

    for item in data:
        # 파일 URL 생성
        file_url = item['download_url']
        file_name = item['name']

        # 파일 다운로드
        response = requests.get(file_url, headers=headers)

        try:
            json_data = json.loads(response.text)  # response를 text로 변환
        except json.JSONDecodeError as e:
            print(f"JSON 디코딩 오류: {str(e)}")
            continue

        cve_number = json_data["CVE_data_meta"]["ID"]
        csv_file_name = os.path.join("CVE_files", f"CVE-{year}-{cve_number}.csv")  # 수정: 확장자 변경
        df = pd.json_normalize(json_data)

        # CSV 파일로 저장
        df.to_csv(csv_file_name, encoding='utf-8-sig') 

except json.decoder.JSONDecodeError:
    # JSON 디코딩 오류 처리
    print("JSON decoding error. Response content:")
    print(response.text)
// 출력되는 데이터가 여기서 나오는 데이터

# 크롬 드라이버 종료
driver.quit()


import requests
import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import StaleElementReferenceException
import pandas as pd
import json

# Chrome 웹 드라이버 실행
s = Service('/opt/homebrew/bin/chromedriver')
driver = webdriver.Chrome(service=s)
driver.maximize_window()
time.sleep(10)

# GitHub API 엔드포인트 및 토큰 설정
base_url = 'https://github.com/CVEProject/cvelistV5/tree/main/cves'
token = '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'

# 다운로드할 연도
year = 2023

# 결과를 저장할 디렉토리 생성
if not os.path.exists('CVE_files'):
    os.makedirs('CVE_files')

year_url = f'{base_url}{year:04d}'
headers = {'Authorization': f'token {token}'}

# 해당 연도의 폴더 내 파일 목록 가져오기
response = requests.get(year_url, headers=headers)

try:
    data = response.json()
    result_df = pd.DataFrame(columns=['File Name', 'Code'])

    for item in data:
        if item['type'] == 'file':
            # 파일 URL 생성
            file_url = item['download_url']
            file_name = item['name']

            # 파일 다운로드
            response = requests.get(file_url, headers=headers)

            try:
                json_data = json.loads(response.text)  # response를 text로
            except json.JSONDecodeError as e:
                print(f"디코딩 오류: {str(e)}")
                continue

            cve_number = json_data["CVE_data_meta"]["ID"]
            csv_file_name = os.path.join("CVE_files", f"CVE-{year}-{cve_number}.csv") 
            df = pd.json_normalize(json_data)

            # CSV 파일로 저장
            df.to_csv("cve.csv", encoding='utf-8-sig') 
            #df.to_csv("cve.csv", encoding='utf-8') 


except json.decoder.JSONDecodeError:
    print("JSON decoding error")
    print(response.text)

driver.quit()




# step 1
#https://github.com/CVEProject/cvelistV5/tree/main/cves/2023
#https://github.com/CVEProject/cvelistV5/tree/main/cves/2023/0xxx
#https://github.com/CVEProject/cvelistV5/blob/main/cves/2023/0xxx/CVE-2023-0001.json
# 이 라인대로 타고 타고 들어가서 크롤링을 해야함

# step 2
#//*[@id="folder-row-1"]/td[2]/div/div/h3/div
#//*[@id="folder-row-5"]/td[2]/div/div/h3/div
# 이 라인대로 0xxx 1xxx 20xxx ... 이렇게 눌러서 크롤링을 해야됨 

#이게 차례대로 0001, 0002, 0003.json 파일의 xpath
#//*[@id=":r2k:"]/span/span
#//*[@id=":r2n:"]/span/span
#//*[@id=":r2q:"]/span/span

# step 3
#//*[@id="repos-sticky-header"]/div[1]/div[2]/div[2]/div[1]/div[1]/span/button/svg
# 이게 파일 다운로드 버튼 XPATH 
# //*[@id="repo-content-pjax-container"]/react-app/div/div/div[1]/div/div/main/div[2]/div/div[1]/div/div[1]/div/div[1]/h2/button[1]/span
# 이게 뒤로가기 버튼


from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
import time
from selenium.common.exceptions import StaleElementReferenceException

from selenium.common.exceptions import NoSuchElementException, TimeoutException

import openpyxl

wb = openpyxl.Workbook()
ws = wb.active

s = Service('/opt/homebrew/bin/chromedriver')
driver = webdriver.Chrome(service=s)


url = 'https://github.com/CVEProject/cvelistV5/tree/main/cves'

driver.get(url)
driver.maximize_window()
driver.implicitly_wait(time_to_wait=5)

year_element = driver.find_element(By.XPATH, '//*[@id="folder-row-25"]/td[2]/div/div/h3/div/a')
year_element.click()
try:
    add_element = driver.find_element(By.XPATH, '//*[@id="folder-row-1"]/td[2]/div/div/h3/div/a')
    add_element.click()
except StaleElementReferenceException:
    # 예외 처리: 요소가 더 이상 사용 가능하지 않을 때 다시 찾기
    add_element = driver.find_element(By.XPATH, '//*[@id="folder-row-1"]/td[2]/div/div/h3/div/a')  # 다시 찾기
    add_element.click()


while True:
    try:
        path = """//*[@id="repo-content-pjax-container"]/react-app/div/div/div[1]/div/div/main/div[2]/div/div[3]"""
        rows = driver.find_elements(By.XPATH,path)
        for i in range(len(rows)):
            row = rows[i]
            # 페이지 소스 가져오기
            page_source = driver.page_source

            # 파일 이름과 내용 엑셀에 추가
            ws.append([f'CVE_data_{i}', page_source])
            print(f"Data saved for CVE_data_{i}")
            
            back_element = '//*[@id="repo-content-pjax-container"]/react-app/div/div/div[1]/div/div/main/div[2]/div/div[1]/div/div[1]/div/div[1]/h2/button[1]/span'
            back_element.click()
                time.sleep(2)
                i += 1
            except NoSuchElementException:
                pass

    except StaleElementReferenceException:
        break


wb.save("cve_data.xlsx")
driver.quit()

  # 페이지가 로드될 때까지 대기


            if response.status_code == 200:
                # 파일 저장
                with open(os.path.join('CVE_files', file_name), 'wb') as file:
                    file.write(response.content)
                print(f'Downloaded: {file_name}')

                time.sleep(10)

                with open(os.path.join('CVE_files', file_name), 'r', encoding='utf-8') as file:
                    file_content = file.read()
            
                # 데이터 프레임에 추가
                result_df = result_df.append({'File Name': file_name, 'Content': file_content}, ignore_index=True)

                #파일로 저장
                result_df.to_excel('CVE_data_2023.xlsx', index=False)



from inspect import getfile
import os
import re
from urllib import request
from urllib.error import HTTPError
from bs4 import BeautifulSoup
import requests

import os
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from bs4 import BeautifulSoup

s = Service('/opt/homebrew/bin/chromedriver')
driver = webdriver.Chrome(service=s)
driver.maximize_window()

github_url = 'https://github.com/CVEProject/cvelistV5/tree/main/cves/2023'
driver.get(github_url)

# 결과를 저장할 디렉토리 생성
output_directory = 'CVE_files_2023'
if not os.path.exists(output_directory):
    os.makedirs(output_directory)

crawled_data = []

# GitHub 사이트를 탐색하면서 파일 다운로드
while True:
    try:
        # 현재 페이지의 HTML 파싱
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        
        file_elements = soup.find_all('a', class_='js-navigation-open link-gray-dark')
        for file_element in file_elements:
            file_name = file_element.get_text()
            file_url = file_element['href']
            
            # 파일의 확장자가 .json인 경우에만 다운로드 수행
            if file_name.endswith('.json'):
                # 파일 다운로드를 위해 해당 파일 페이지로 이동
                driver.get('https://github.com' + file_url)
                time.sleep(2)  # 페이지 로딩을 기다림
                
                # 파일 다운로드 버튼을 찾아 클릭
                try:
                    download_button = driver.find_element(By.XPATH, '//*[@id="repo-content-pjax-container"]/react-app/div/div/div[1]/div/div/main/div[2]/div/div[1]/div/div[1]/div/div[1]/h2/button[1]/span')
                    download_button.click()
                except NoSuchElementException:
                    print(f"파일 다운로드 버튼을 찾을 수 없습니다: {file_name}")
                
                # 다운로드한 파일을 디렉토리에 이동 및 이름 변경
                time.sleep(5)  
                downloaded_file_name = max([os.path.join(output_directory, f) for f in os.listdir(output_directory)], key=os.path.getctime)
                new_file_name = os.path.join(output_directory, file_name)
                os.rename(downloaded_file_name, new_file_name)
                
                print(f"다운로드 및 저장 완료: {new_file_name}")
                
                # 결과 DataFrame에 추가
                crawled_data.append({'File Name': file_name, 'Content': new_file_name})
        
        more_button = driver.find_element(By.PARTIAL_LINK_TEXT, 'Load more')
        more_button.click()
        time.sleep(2)  
        
    except NoSuchElementException:
        print("파일없다")
        break
    except TimeoutException:
        print("타임아웃")

result_df = pd.DataFrame(crawled_data)


# 결과를 엑셀 파일로 저장
result_excel_file = os.path.join(output_directory, 'CVE_files_2023.xlsx')
result_df.to_excel(result_excel_file, index=False)
print(f"저장 완료{result_excel_file}")

# 크롬 드라이버 종료
driver.quit()

import os
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from bs4 import BeautifulSoup
import openpyxl

s = Service('/opt/homebrew/bin/chromedriver')
driver = webdriver.Chrome(service=s)
driver.maximize_window()

github_url = 'https://github.com/CVEProject/cvelistV5/tree/main/cves/2023'
driver.get(github_url)

# 결과 저장할 디렉토리 생성
output_directory = 'CVE_files_2023'
if not os.path.exists(output_directory):
    os.makedirs(output_directory)

crawled_data = []  

# GitHub 사이트를 탐색하면서 파일 다운로드
while True:
    try:
        # 현재 페이지의 HTML 파싱
        page_source = "https://github.com/CVEProject/cvelistV5/tree/main/cves"
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        
        file_elements = soup.find_all('a', class_='js-navigation-open link-gray-dark')
        for file_element in file_elements:
            file_name = file_element.get_text()
            file_url = file_element['href']
            
            # 파일의 확장자가 .json인 경우에만 다운로드 수행
            if file_name.endswith('.json'):
                # 파일 다운로드를 위해 해당 파일 페이지로 이동
                driver.get('https://github.com' + file_url)
                time.sleep(2)  
                
                # 파일 다운로드 버튼을 찾아 클릭
                try:
                    download_button = driver.find_element(By.XPATH, '//*[@id="repo-content-pjax-container"]/react-app/div/div/div[1]/div/div/main/div/section[2]/div/div[1]/div[1]/div[2]/div/div/div[3]/div[2]/div[1]/span/a')
                    download_button.click()
                except NoSuchElementException:
                    print(f"다운로드 버튼 없음: {file_name}")
                
                # 다운로드한 파일을 디렉토리에 이동 및 이름 변경
                time.sleep(5)  # 다운로드가 완료될 때까지 대기
                downloaded_file_name = max([os.path.join(output_directory, f) for f in os.listdir(output_directory)], key=os.path.getctime)
                new_file_name = os.path.join(output_directory, file_name)
                os.rename(downloaded_file_name, new_file_name)
                
                print(f"다운로드 및 저장 완료: {new_file_name}")
                
                # 결과 리스트에 추가
                crawled_data.append({'File Name': file_name, 'Content': new_file_name})
        
        more_button = driver.find_element(By.PARTIAL_LINK_TEXT, 'Load more')
        more_button.click()
        time.sleep(2)  # 페이지 로딩을 기다림
        
    except NoSuchElementException:
        print("파일없다")
        break
    except TimeoutException:
        print("타임아웃")

result_df = pd.DataFrame(crawled_data)

# 결과를 엑셀 파일로 저장
result_excel_file = os.path.join(output_directory, 'CVE_files_2023.xlsx')
result_df.to_excel(result_excel_file, index=False)
print(f"{result_excel_file}에 저장 완료")


# 크롬 드라이버 종료
driver.quit()

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import openpyxl


wb = openpyxl.Workbook()
ws = wb.active


s = Service('/opt/homebrew/bin/chromedriver')
driver = webdriver.Chrome(service=s)

url = 'https://github.com/CVEProject/cvelistV5/tree/main/cves'

driver.get(url)
driver.maximize_window()
driver.implicitly_wait(time_to_wait=5)

year_element = driver.find_element(By.XPATH, '//*[@id="folder-row-25"]/td[2]/div/div/h3/div/a')
year_element.click()
try:
    add_element = driver.find_element(By.XPATH, '//*[@id="folder-row-1"]/td[2]/div/div/h3/div/a')
    add_element.click()
except StaleElementReferenceException:
    # 예외 처리: 요소가 더 이상 사용 가능하지 않을 때 다시 찾기
    add_element = driver.find_element(By.XPATH, '//*[@id="folder-row-1"]/td[2]/div/div/h3/div/a')  # 다시 찾기
    add_element.click()

i = 1
while True:
    try:
        # 페이지 소스 가져오기
        page_source = driver.page_source

        # 파일 이름과 내용 엑셀에 추가
        ws.append([f'CVE_data_{i}', page_source])
        print(f"Data saved for CVE_data_{i}")
        
        back_element_xpath = '//*[@id="repo-content-pjax-container"]/react-app/div/div/div[1]/div/div/main/div[2]/div/div[1]/div/div[1]/div/div[1]/h2/button[1]/span'
        try:
            next_element = driver.find_element(By.XPATH, back_element_xpath)
            next_element.click()
            time.sleep(2)
            i += 1
        except NoSuchElementException:
            break  # 더 이상 다음 페이지가 없을 때 반복문 종료

    except StaleElementReferenceException:
        break

wb.save("cve_data.xlsx")
driver.quit()


import os
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from bs4 import BeautifulSoup
import openpyxl

s = Service('/opt/homebrew/bin/chromedriver')
driver = webdriver.Chrome(service=s)
driver.maximize_window()

github_url = 'https://github.com/CVEProject/cvelistV5/tree/main/cves/2023'
driver.get(github_url)

# 결과 저장할 디렉토리 생성
output_directory = 'CVE_files_2023'
if not os.path.exists(output_directory):
    os.makedirs(output_directory)

crawled_data = []  

# GitHub 사이트를 탐색하면서 파일 다운로드
while True:
    try:
        # 현재 페이지의 HTML 파싱
        page_source = "https://github.com/CVEProject/cvelistV5/tree/main/cves"
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        
        file_elements = soup.find_all('a', class_='js-navigation-open link-gray-dark')
        for file_element in file_elements:
            file_name = file_element.get_text()
            file_url = file_element['href']
            
            # 파일의 확장자가 .json인 경우에만 다운로드 수행
            if file_name.endswith('.json'):
                # 파일 다운로드를 위해 해당 파일 페이지로 이동
                driver.get('https://github.com' + file_url)
                time.sleep(2)  # 페이지 로딩을 기다림
                
                # 파일 다운로드 버튼을 찾아 클릭
                try:
                    download_button = driver.find_element(By.XPATH, '//*[@id="repo-content-pjax-container"]/react-app/div/div/div[1]/div/div/main/div/section[2]/div/div[1]/div[1]/div[2]/div/div/div[3]/div[2]/div[1]/span/a')
                    download_button.click()
                except NoSuchElementException:
                    print(f"다운로드 버튼 없음: {file_name}")
                
                # 다운로드한 파일을 디렉토리에 이동 및 이름 변경
                time.sleep(5)  # 다운로드가 완료될 때까지 대기
                downloaded_file_name = max([os.path.join(output_directory, f) for f in os.listdir(output_directory)], key=os.path.getctime)
                new_file_name = os.path.join(output_directory, file_name)
                os.rename(downloaded_file_name, new_file_name)
                
                print(f"다운로드 및 저장 완료: {new_file_name}")
                
                # 결과 리스트에 추가
                crawled_data.append({'File Name': file_name, 'Content': new_file_name})
        
        # "Load more" 버튼이 있는지 확인하여 클릭
        load_more_button = driver.find_element(By.PARTIAL_LINK_TEXT, 'Load more')
        load_more_button.click()
        time.sleep(2)  # 페이지 로딩을 기다림
        
    except NoSuchElementException:
        print("파일없다")
        break
    except TimeoutException:
        print("타임아웃")

result_df = pd.DataFrame(crawled_data)

# 결과를 엑셀 파일로 저장
result_excel_file = os.path.join(output_directory, 'CVE_files_2023.xlsx')
result_df.to_excel(result_excel_file, index=False)
print(f"{result_excel_file}에 저장 완료")


# 크롬 드라이버 종료
driver.quit()


import requests
import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import StaleElementReferenceException
import pandas as pd
import json

# Chrome 웹 드라이버를 실행
s = Service('/opt/homebrew/bin/chromedriver')
driver = webdriver.Chrome(service=s)
driver.maximize_window()
time.sleep(10)

# GitHub API 엔드포인트 및 토큰 설정
base_url = 'https://github.com/CVEProject/cvelistV5/tree/main/cves'
token = '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'

# 다운로드할 연도 설정
year = 2023

# 결과를 저장할 디렉토리 생성
if not os.path.exists('CVE_files'):
    os.makedirs('CVE_files')

year_url = f'{base_url}{year:04d}'
headers = {'Authorization': f'token {token}'}
    
# 해당 연도의 폴더 내 파일 목록 가져오기
response = requests.get(year_url, headers=headers)


try:
    data = response.json()
    result_df = pd.DataFrame(columns=['File Name', 'Content'])

    for item in data:
        if item['type'] == 'file':
            # 파일 URL 생성
            file_url = item['download_url']
            file_name = item['name']
            
            # 파일 다운로드
            response = requests.get(file_url, headers=headers)
            
            try:
                json_data = json.loads(response)
            except json.JSONDecodeError as e:
                print(f"JSON 디코딩 오류: {str(e)}")
                continue
            
            cve_number = json_data["CVE_data_meta"]["ID"]
            excel_file_name = os.path.join("CVE_files", f"CVE-{year}-{cve_number}.xlsx")
            df = pd.json_normalize(json_data)
            df.to_excel(excel_file_name, index=False)

except json.decoder.JSONDecodeError:
    # JSON 디코딩 오류 처리
    print("JSON decoding error. Response content:")
    print(response.text)


# 크롬 드라이버 종료
driver.quit()

import requests
import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import StaleElementReferenceException
import pandas as pd
import json

# Chrome 웹 드라이버를 실행
s = Service('/opt/homebrew/bin/chromedriver')
driver = webdriver.Chrome(service=s)
driver.maximize_window()
time.sleep(10)

# GitHub API 엔드포인트 및 토큰 설정
base_url = 'https://github.com/CVEProject/cvelistV5/tree/main/cves'
token = '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'

# 다운로드할 연도 설정
year = 2023

# 결과를 저장할 디렉토리 생성
if not os.path.exists('CVE_files'):
    os.makedirs('CVE_files')

year_url = f'{base_url}{year:04d}'
headers = {'Authorization': f'token {token}'}

# 해당 연도의 폴더 내 파일 목록 가져오기
response = requests.get(year_url, headers=headers)

try:
    data = response.json()
    result_df = pd.DataFrame(columns=['File Name', 'Content'])

    for item in data:
        if item['type'] == 'file':
            # 파일 URL 생성
            file_url = item['download_url']
            file_name = item['name']

            # 파일 다운로드
            response = requests.get(file_url, headers=headers)

            try:
                json_data = json.loads(response.text)  # 수정: response를 text로 변환
            except json.JSONDecodeError as e:
                print(f"JSON 디코딩 오류: {str(e)}")
                continue

            cve_number = json_data["CVE_data_meta"]["ID"]
            csv_file_name = os.path.join("CVE_files", f"CVE-{year}-{cve_number}.csv")  # 수정: 확장자 변경
            df = pd.json_normalize(json_data)

            # CSV 파일로 저장
            df.to_csv("cve.csv", encoding='utf-8-sig') 

except json.decoder.JSONDecodeError:
    # JSON 디코딩 오류 처리
    print("JSON decoding error. Response content:")
    print(response.text)

# 크롬 드라이버 종료
driver.quit()


import requests
import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import StaleElementReferenceException
import pandas as pd
import json

# Chrome 웹 드라이버를 실행
s = Service('/opt/homebrew/bin/chromedriver')
driver = webdriver.Chrome(service=s)
driver.maximize_window()
time.sleep(10)

# GitHub API 엔드포인트 및 토큰 설정
base_url = 'https://github.com/CVEProject/cvelistV5/tree/main/cves'
token = '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'

# 다운로드할 연도 설정
year = 2023

# 결과를 저장할 디렉토리 생성
if not os.path.exists('CVE'):
    os.makedirs('CVE')

year = f'{base_url}/{year:04d}'
headers = {'Authorization': f'token {token}'}

# 해당 연도의 폴더 내 파일 목록 가져오기
response = requests.get(year, headers=headers)

# 파일 정보를 저장할 리스트 생성
file_data_list = []

try:
    data = response.json()

    for item in data:
        if item['type'] == 'file':
            # 파일 URL 생성
            file_url = item['download_url']
            file_name = item['name']

            # 파일 다운로드
            response = requests.get(file_url, headers=headers)

            try:
                json_data = json.loads(response.text) 
                print("성공적으로 디코딩")
                cve_number = json_data["CVE_data_meta"]["ID"]
                file_data_list.append({'File Name': file_name, 'Content': json_data})

            except json.JSONDecodeError as e:
                print(f"JSON 디코딩 오류: {str(e)}")
                continue
            
except json.decoder.JSONDecodeError:
    # JSON 디코딩 오류 처리
    print("JSON decoding error. Response content:")
    print(response.text)

# 크롬 드라이버 종료
driver.quit()

# 리스트 --> 데이터프레임
result_df = pd.DataFrame(file_data_list)

# CSV 파일로 저장
result_df.to_csv("cve_data.csv", index=False, encoding='utf-8-sig')
'''
